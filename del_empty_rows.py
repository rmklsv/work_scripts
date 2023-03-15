import openpyxl

# Load the destination workbook and worksheet
dest_wb = openpyxl.load_workbook("path_to_the_file/dest_file.xlsx")
dest_ws = dest_wb.active

# Find the last row of data in column B
last_row = dest_ws.max_row

# Delete empty rows below the last row of data in column B
for row in range(last_row + 1, dest_ws.max_row + 1):
    if all(cell.value is None for cell in dest_ws[row]):
        dest_ws.delete_rows(row)

# Save the changes to the destination workbook
dest_wb.save("path_to_the_file/dest_file.xlsx")