import shutil
import openpyxl
from openpyxl.utils import range_boundaries

# Create backup
s_path = r"path_to_the_file\source_file.xlsx"
d_path = r"path_to_the_file\source_file_backup.xlsx"
shutil.copy(s_path, d_path)
print("Backup file created")

source_wb = openpyxl.load_workbook("path_to_the_file/source_file.xlsx", data_only=True)
source_ws = source_wb.active

dest_wb = openpyxl.load_workbook("path_to_the_file/dest_file.xlsx", data_only=True)
dest_ws = dest_wb.active

row_numbers = input("Enter the rows "
                    "to update (e.g. 1,2,3): ").split(',')

for row_number in row_numbers:
    # Define the range of necessary columns
    range_string = f"A{row_number}:B{row_number};D{row_number};J{row_number}"
    # Define the cell with vendor code
    vendor_cell = source_ws[f"B{row_number}"]
    # Defines trigger for the next part of code
    found_match = False
    for dest_cell in dest_ws["B"]:
        # Check vendor code if item already exist in nomenclature
        if vendor_cell.value == dest_cell.value:
            row_index = dest_cell.row
            # Update the values
            dest_ws[f"C{row_index}"].value = source_ws[f"D{row_number}"].value
            dest_ws[f"D{row_index}"].value = source_ws[f"J{row_number}"].value
            # Update trigger for the next part of code
            found_match = True
            # Vendor code is unique, that's why only 1 match possible
            break
    
    # Check trigger
    if not found_match:
        # Parse the range string into separate cell ranges
        cell_ranges = range_string.split(';')
        
        values = []

        # Loop through each cell range and extract the cells
        for cell_range in cell_ranges:
            (min_col, min_row, 
            max_col, max_row) = range_boundaries(cell_range)
            for row in source_ws.iter_rows(min_row=min_row, 
                                        max_row=max_row, 
                                        min_col=min_col, 
                                        max_col=max_col):
                for cell in row:
                    values.append(cell.value)

        # Append the values to the destination worksheet
        dest_ws.append(values)

dest_wb.save("path_to_the_file/dest_file.xlsx")
print("Sucessfully updated")